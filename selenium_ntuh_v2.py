from selenium import webdriver
from bs4 import BeautifulSoup
from selenium.webdriver.support.ui import Select
import urllib.request
import pytesseract
from PIL import Image
import PIL.ImageOps
from PIL import ImageDraw
from openpyxl import load_workbook
import time
import re
import sys
import shutil

ELE = []
offset = 180
def initTable(threshold=180):   #  阈值的设置对图片除燥意义重大，有算法可以 计算，不知道算法的话，手动尝试
    table = []
    for i in range(256):
        if i < threshold:
            table.append(0)
        else:
            table.append(1)
    return table

def depoint(img):   #input: gray image
     pixdata = img.load()
     w,h = img.size
     for y in range(1,h-1):
         for x in range(1,w-1):
             count = 0
             if pixdata[x,y-1] > offset:
                 count = count + 1
             if pixdata[x,y+1] > offset:
                 count = count + 1
             if pixdata[x-1,y] > offset:
                 count = count + 1
             if pixdata[x+1,y] > offset:
                 count = count + 1
             if count > 2:
                 pixdata[x,y] = 255
     return img

def analyze():
	im = Image.open('code.gif')
	#图片的处理过程
	im = im.convert('L')
	depoint(im)
	im.save("depoint.jpg")
	code = pytesseract.image_to_string(im, config='-psm 7 -c tessedit_char_whitelist=0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ')
	return code

def download_img(browser):
	# get the image source
	img_src = browser.find_element_by_id("UclQueryInput_imgVlid").get_attribute("src")
	#print(img_src)
	# download the image
	urllib.request.urlretrieve(img_src, "code.gif")

def fill_patient(browser, idno, bir_y, bir_m, bir_d, codenum):
	print(idno)
	python_button = browser.find_elements_by_xpath("//input[@name='UclQueryInput:radInputNum' and @value='1']")[0]
	python_button.click()

	text_area = browser.find_element_by_id('UclQueryInput_txtIdno')
	text_area.send_keys(idno)
	
	sel_y = Select(browser.find_element_by_id('UclQueryInput_ddlBirthYear'))
	sel_y.select_by_value(bir_y)
	sel_m = Select(browser.find_element_by_id('UclQueryInput_ddlBirthMonth'))
	sel_m.select_by_value(bir_m)
	sel_d = Select(browser.find_element_by_id('UclQueryInput_ddlBirthDay'))
	sel_d.select_by_value(bir_d)
	
	text_codenum = browser.find_element_by_id('UclQueryInput_txtVerifyCode')
	text_codenum.send_keys(codenum)

def query_patient(sheet, i, browser, idno, bir_y, bir_m, bir_d):

	recheck = True
	while recheck == True:
		download_img(browser)
		codenum = analyze()
		codenum.strip().lstrip().rstrip(',').strip(' ').lstrip('‘').rstrip('‘')
		codenum = codenum.replace(" ", "")

		if len(codenum) == 6:
			recheck = False
		print(codenum)

	fill_patient(browser, idno, bir_y, bir_m, bir_d, codenum)
	browser.find_element_by_id('UclQueryInput_btnQuery').click()
	try:
		obj = browser.switch_to.alert
		#browser.close()
		#Retrieve the message on the Alert window
		msg=obj.text
		print ("Alert shows following message: "+ msg )
		if msg == "驗證碼輸入錯誤，請重新輸入" :
			print("Strings are equal with text : ", msg)
			obj.accept()
			browser.get(browser.current_url)
			#dest = shutil.copyfile("afterbi.jpg", "afterbi_2.jpg")
			dest = shutil.copyfile("depoint.jpg", "depoint_2.jpg")
			return False
		elif msg == "病人查無未來掛號資料" :
			obj.accept()
			#browser.forward()
			browser.get(browser.current_url)
			return True
		elif msg == "請輸入正確身分證號..." :
			obj.accept()
			browser.get(browser.current_url)
			return True
		elif msg == "請輸入正確出生年月日" :
			obj.accept()
			browser.get(browser.current_url)
			return True
		else :
			print(msg)
			print ("Strings are not equal")
			return False
	except:
		try:
			medical_recordnum = browser.find_element_by_id('UclQueryInput_txtPassword')
			print("需要輸入病歷號碼")
			browser.get(browser.current_url)
			return True
		except:
			table_html = browser.find_element_by_id('GridViewPatientAdminServiceList').text
			sheet.cell(row=i,column=3).value = table_html
			sheet.row_dimensions[i].height = 50
			#print(table_html)
			print("no alert to accept")
			browser.get(browser.current_url)
			return True
def has_special_char(string): 
  
	# Make own character set and pass  
	# this as argument in compile method 
	regex = re.compile('▲')
	# Pass the string in search  
	# method of regex object.     
	if(regex.search(string) == None): 
		print("String is accepted")
		return False
	else: 
		print("String is not accepted.")
		return True

def read_xlxs(browser):
	wb = load_workbook('data.xlsx')
	sheet = wb.active
	sheet.column_dimensions["c"].width = 100
	max_row=sheet.max_row
	for i in range(1,max_row+1):

		bir = sheet.cell(row=i,column=1).value.lstrip('男').lstrip('女')
		bir_array = bir.split("/")
		bir_y = bir_array[0].lstrip('0')
		bir_y = str(int(bir_y)+1911)
		bir_m = bir_array[1] 
		bir_d = bir_array[2] 
		idno = sheet.cell(row=i,column=2).value
		if(has_special_char(idno)):
			continue
		query_has_record = False
		failcount = 0
		while query_has_record == False:
			query_has_record = query_patient(sheet, i, browser, idno, bir_y, bir_m, bir_d)
			failcount += 1
			if(failcount > 3):
                                browser.refresh()
                                failcount = 0
			else:
                                print(failcount)

	wb.close()    #关闭工作薄
	wb.save('data.xlsx')

def main():
	browser = webdriver.Chrome()
	browser.get('https://reg.ntuh.gov.tw/webadministration/Query.aspx')
	

	read_xlxs(browser)

if __name__ == '__main__':

	main()
